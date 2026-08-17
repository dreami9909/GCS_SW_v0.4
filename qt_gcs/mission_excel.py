from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from .site_store import MissionZone, SiteStore


HEADER_ALIASES = {
    "TYPE": "TYPE",
    "종류": "TYPE",
    "ID": "ID",
    "CODE": "ID",
    "코드": "ID",
    "LATITUDE": "LATITUDE",
    "LAT": "LATITUDE",
    "위도": "LATITUDE",
    "LONGITUDE": "LONGITUDE",
    "LONG": "LONGITUDE",
    "LON": "LONGITUDE",
    "LNG": "LONGITUDE",
    "경도": "LONGITUDE",
    "ALTITUDE_M": "ALTITUDE_M",
    "ALTITUDE": "ALTITUDE_M",
    "ALT": "ALTITUDE_M",
    "고도": "ALTITUDE_M",
    "ORDER": "ORDER",
    "SEQUENCE": "ORDER",
    "SEQ": "ORDER",
    "VERTEX_ORDER": "ORDER",
    "순서": "ORDER",
    "NOTES": "NOTES",
    "비고": "NOTES",
}

TYPE_ALIASES = {
    "GCS": "GCS",
    "지상통제소": "GCS",
    "RDR": "RDR",
    "RADAR": "RDR",
    "레이다": "RDR",
    "레이더": "RDR",
    "LC": "LC",
    "LAUNCHER": "LC",
    "발사대": "LC",
    "WAYPOINT": "WAYPOINT",
    "WP": "WAYPOINT",
    "웨이포인트": "WAYPOINT",
    "RETURN": "RETURN",
    "RTB": "RETURN",
    "복귀지점": "RETURN",
    "SAFE": "SAFE_ZONE",
    "SAFE_ZONE": "SAFE_ZONE",
    "안전지대": "SAFE_ZONE",
}


class MissionExcelError(ValueError):
    pass


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _number(value: Any, *, row: int, field: str, default: float | None = None) -> float:
    if value is None or _text(value) == "":
        if default is not None:
            return default
        raise MissionExcelError(f"{row}행 {field} 값이 비어 있습니다.")
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise MissionExcelError(
            f"{row}행 {field} 값은 숫자여야 합니다: {value!r}"
        ) from error


def load_mission_workbook(path: str | Path) -> SiteStore:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise MissionExcelError(
            "Excel 읽기 모듈이 없습니다. requirements.txt를 다시 설치하십시오."
        ) from error

    try:
        workbook = load_workbook(
            filename=Path(path),
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise MissionExcelError(f"Excel 파일을 열 수 없습니다: {error}") from error

    try:
        worksheet = (
            workbook["Mission"]
            if "Mission" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            raise MissionExcelError("Excel 시트가 비어 있습니다.")

        headers: dict[str, int] = {}
        for column, value in enumerate(raw_headers):
            normalized = HEADER_ALIASES.get(_text(value).upper())
            if normalized and normalized not in headers:
                headers[normalized] = column
        missing = {"TYPE", "LATITUDE", "LONGITUDE"} - set(headers)
        if missing:
            raise MissionExcelError(
                "필수 열이 없습니다: " + ", ".join(sorted(missing))
            )

        def cell(values: tuple[Any, ...], name: str) -> Any:
            index = headers.get(name)
            return values[index] if index is not None and index < len(values) else None

        parsed_rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            raw_type = _text(cell(values, "TYPE"))
            item_type = TYPE_ALIASES.get(raw_type.upper())
            if item_type is None:
                raise MissionExcelError(
                    f"{row_number}행 TYPE을 지원하지 않습니다: {raw_type!r}"
                )
            parsed_rows.append(
                {
                    "row": row_number,
                    "type": item_type,
                    "id": _text(cell(values, "ID")),
                    "latitude": _number(
                        cell(values, "LATITUDE"),
                        row=row_number,
                        field="LATITUDE",
                    ),
                    "longitude": _number(
                        cell(values, "LONGITUDE"),
                        row=row_number,
                        field="LONGITUDE",
                    ),
                    "altitude_m": _number(
                        cell(values, "ALTITUDE_M"),
                        row=row_number,
                        field="ALTITUDE_M",
                        default=0.0,
                    ),
                    "order": _number(
                        cell(values, "ORDER"),
                        row=row_number,
                        field="ORDER",
                        default=float(row_number),
                    ),
                }
            )
    finally:
        workbook.close()

    if not parsed_rows:
        raise MissionExcelError("장입할 임무 데이터가 없습니다.")

    imported = SiteStore()
    waypoint_rows = sorted(
        (row for row in parsed_rows if row["type"] == "WAYPOINT"),
        key=lambda row: (row["order"], row["row"]),
    )
    safe_zone_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)

    try:
        for row in parsed_rows:
            item_type = row["type"]
            if item_type in SiteStore.ALLOWED_CODES:
                imported.set_site(
                    item_type,
                    row["latitude"],
                    row["longitude"],
                    row["altitude_m"],
                )
            elif item_type == "RETURN":
                # Kept as a legacy workbook token; current missions return to
                # the center of the shared SAFE zone instead.
                continue
            elif item_type == "SAFE_ZONE":
                safe_zone_rows[row["id"] or "SAFE01"].append(row)

        for row in waypoint_rows:
            imported.add_waypoint(
                row["latitude"],
                row["longitude"],
                row["altitude_m"] or 60.0,
            )

        for sequence, (zone_id, vertices) in enumerate(
            sorted(safe_zone_rows.items()),
            start=1,
        ):
            ordered = sorted(vertices, key=lambda row: (row["order"], row["row"]))
            if len(ordered) < 3:
                raise MissionExcelError(
                    f"안전지대 {zone_id!r}에는 경계점이 3개 이상 필요합니다."
                )
            points = [
                imported._validated_point(
                    row["latitude"],
                    row["longitude"],
                    row["altitude_m"],
                )
                for row in ordered
            ]
            code = zone_id.upper().replace(" ", "_")
            if not code.startswith("SAFE"):
                code = f"SAFE{sequence:02d}"
            imported.zones.append(
                MissionZone(
                    code=code,
                    label=f"Safe Zone {sequence}",
                    zone_type="SAFE",
                    vertices=points,
                )
            )
    except ValueError as error:
        if isinstance(error, MissionExcelError):
            raise
        raise MissionExcelError(str(error)) from error

    return imported
